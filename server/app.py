"""
app.py — Flask-сервер системы учёта посещаемости студентов.

Запуск:  python app.py
Адрес:   http://localhost:5000
Доступ:  admin/admin  |  teacher/teacher
"""

import os, re, uuid, shutil, threading
from io import BytesIO
from datetime import datetime, date

from flask import Flask, request, jsonify, send_from_directory, abort
from flask_cors import CORS
from flask_jwt_extended import (
    JWTManager, create_access_token,
    jwt_required, get_jwt_identity, get_jwt,
)
from PIL import Image

import config
from models import db, User, Group, Subject, Lesson, Student, StudentPhoto, Attendance
from deepface_service import init_service, get_service

# Хранилище фоновых задач распознавания: job_id -> {status, result?, message?}
_jobs      = {}
_jobs_lock = threading.Lock()


def create_app():
    app = Flask(__name__, static_folder=config.WEB_DIR, static_url_path='')
    app.config.update(
        SECRET_KEY                     = config.SECRET_KEY,
        JWT_SECRET_KEY                 = config.JWT_SECRET_KEY,
        JWT_ACCESS_TOKEN_EXPIRES       = config.JWT_TOKEN_LIFETIME,
        SQLALCHEMY_DATABASE_URI        = config.DATABASE_URI,
        SQLALCHEMY_TRACK_MODIFICATIONS = False,
        MAX_CONTENT_LENGTH             = config.MAX_CONTENT_LENGTH,
    )
    CORS(app); db.init_app(app); JWTManager(app)
    with app.app_context():
        db.create_all()
        _seed_users()
        init_service(config)
    register_routes(app)
    return app


def register_routes(app):
    # Реальный объект приложения для передачи в фоновые потоки.
    # current_app — прокси, недоступный вне контекста HTTP-запроса.
    flask_app = app

    #  Статика 

    @app.route('/')
    def index(): return send_from_directory(config.WEB_DIR, 'index.html')

    @app.route('/admin.html')
    def admin_page(): return send_from_directory(config.WEB_DIR, 'admin.html')

    @app.route('/teacher.html')
    def teacher_page(): return send_from_directory(config.WEB_DIR, 'teacher.html')

    @app.route('/api/photo/<int:sid>/<fn>')
    @jwt_required()
    def serve_photo(sid, fn):
        return send_from_directory(os.path.join(config.PHOTOS_DIR, str(sid)), fn)

    #  Авторизация 

    @app.route('/api/auth/login', methods=['POST'])
    def login():
        """Аутентификация. {username, password} → {access_token, user}."""
        d  = request.get_json(silent=True) or {}
        u  = User.query.filter_by(username=d.get('username','').strip()).first()
        if not u or not u.check_password(d.get('password','')):
            return err('Неверный логин или пароль', 401)
        u.last_login = datetime.utcnow(); db.session.commit()
        token = create_access_token(identity=str(u.id),
            additional_claims={'role': u.role, 'username': u.username})
        return jsonify({'access_token': token, 'user': u.to_dict()})

    @app.route('/api/auth/me')
    @jwt_required()
    def get_me():
        u = User.query.get(int(get_jwt_identity()))
        return jsonify(u.to_dict()) if u else err('Не найден', 404)

    #  Пользователи 

    @app.route('/api/users')
    @jwt_required()
    def users_list():
        admin_only()
        return jsonify([u.to_dict() for u in User.query.order_by(User.full_name).all()])

    @app.route('/api/users', methods=['POST'])
    @jwt_required()
    def users_create():
        admin_only()
        d = request.get_json(silent=True) or {}
        un, pw, role = d.get('username','').strip(), d.get('password',''), d.get('role','teacher')
        if not un or not pw: return err('Логин и пароль обязательны')
        if len(pw) < 4: return err('Пароль: минимум 4 символа')
        if role not in ('admin','teacher'): return err('Роль: admin или teacher')
        if User.query.filter_by(username=un).first(): return err(f'Логин "{un}" занят')
        try:
            u = User(username=un, role=role, full_name=d.get('full_name','').strip() or None)
            u.set_password(pw); db.session.add(u); db.session.commit()
            return jsonify(u.to_dict()), 201
        except Exception as e: db.session.rollback(); return err(str(e), 500)

    @app.route('/api/users/<int:uid>', methods=['DELETE'])
    @jwt_required()
    def users_delete(uid):
        admin_only()
        if uid == int(get_jwt_identity()): return err('Нельзя удалить себя')
        u = User.query.get_or_404(uid)
        try: db.session.delete(u); db.session.commit(); return jsonify({'ok':True})
        except Exception as e: db.session.rollback(); return err(str(e), 500)

    #  Группы 

    @app.route('/api/groups')
    @jwt_required()
    def groups_list():
        return jsonify([g.to_dict() for g in Group.query.order_by(Group.name).all()])

    @app.route('/api/groups', methods=['POST'])
    @jwt_required()
    def groups_create():
        """Создаёт группу. {name*, course, faculty, specialty}"""
        admin_only()
        d = request.get_json(silent=True) or {}
        name = d.get('name','').strip()
        if not name: return err('Название обязательно')
        if Group.query.filter_by(name=name).first(): return err(f'Группа "{name}" существует')
        try:
            g = Group(name=name, course=d.get('course'),
                      faculty=d.get('faculty','').strip() or None,
                      specialty=d.get('specialty','').strip() or None)
            db.session.add(g); db.session.commit(); return jsonify(g.to_dict()), 201
        except Exception as e: db.session.rollback(); return err(str(e), 500)

    @app.route('/api/groups/<int:gid>', methods=['PUT'])
    @jwt_required()
    def groups_update(gid):
        """Изменяет группу. Все поля необязательны."""
        admin_only()
        g = Group.query.get_or_404(gid)
        d = request.get_json(silent=True) or {}
        if 'name'      in d: g.name      = d['name'].strip()      or g.name
        if 'course'    in d: g.course    = d['course']
        if 'faculty'   in d: g.faculty   = d['faculty'].strip()   or None
        if 'specialty' in d: g.specialty = d['specialty'].strip() or None
        try: db.session.commit(); return jsonify(g.to_dict())
        except Exception as e: db.session.rollback(); return err(str(e), 500)

    @app.route('/api/groups/<int:gid>', methods=['DELETE'])
    @jwt_required()
    def groups_delete(gid):
        admin_only()
        g = Group.query.get_or_404(gid)
        try: db.session.delete(g); db.session.commit(); return jsonify({'ok':True})
        except Exception as e: db.session.rollback(); return err(str(e), 500)

    #  Дисциплины 

    @app.route('/api/subjects')
    @jwt_required()
    def subjects_list():
        return jsonify([s.to_dict() for s in Subject.query.order_by(Subject.name).all()])

    @app.route('/api/subjects', methods=['POST'])
    @jwt_required()
    def subjects_create():
        admin_only()
        d = request.get_json(silent=True) or {}
        name = d.get('name','').strip()
        if not name: return err('Название обязательно')
        try:
            s = Subject(name=name, code=d.get('code','').strip() or None,
                        description=d.get('description','').strip() or None)
            db.session.add(s); db.session.commit(); return jsonify(s.to_dict()), 201
        except Exception as e: db.session.rollback(); return err(str(e), 500)

    @app.route('/api/subjects/<int:sid>', methods=['PUT'])
    @jwt_required()
    def subjects_update(sid):
        """Изменяет дисциплину. {name, code, description} — все необязательны."""
        admin_only()
        s = Subject.query.get_or_404(sid)
        d = request.get_json(silent=True) or {}
        if 'name'        in d: s.name        = d['name'].strip()        or s.name
        if 'code'        in d: s.code        = d['code'].strip()        or None
        if 'description' in d: s.description = d['description'].strip() or None
        try: db.session.commit(); return jsonify(s.to_dict())
        except Exception as e: db.session.rollback(); return err(str(e), 500)

    @app.route('/api/subjects/<int:sid>', methods=['DELETE'])
    @jwt_required()
    def subjects_delete(sid):
        admin_only()
        s = Subject.query.get_or_404(sid)
        try: db.session.delete(s); db.session.commit(); return jsonify({'ok':True})
        except Exception as e: db.session.rollback(); return err(str(e), 500)

    # Занятия

    @app.route('/api/lessons')
    @jwt_required()
    def lessons_list():
        """Список занятий. Query: date, group_id, teacher_id."""
        q = Lesson.query
        ds = request.args.get('date','').strip()
        gid = request.args.get('group_id', type=int)
        tid = request.args.get('teacher_id', type=int)
        if ds:
            try: q = q.filter(Lesson.lesson_date == date.fromisoformat(ds))
            except ValueError: return err('Формат даты: YYYY-MM-DD')
        if gid: q = q.filter(Lesson.group_id   == gid)
        if tid: q = q.filter(Lesson.teacher_id == tid)
        return jsonify([l.to_dict() for l in
                        q.order_by(Lesson.lesson_date.desc(), Lesson.time_start).all()])

    @app.route('/api/lessons/current')
    @jwt_required()
    def lessons_current():
        """
        Текущее занятие группы по системному времени.
        Приоритет: идёт сейчас → следующее → последнее за сегодня.
        Query: group_id*.
        """
        gid = request.args.get('group_id', type=int)
        if not gid: return err('group_id обязателен')
        today, now = date.today(), datetime.now().strftime('%H:%M')
        lessons = (Lesson.query.filter_by(group_id=gid, lesson_date=today)
                   .order_by(Lesson.time_start).all())
        if not lessons: return jsonify({'lesson': None})
        for l in lessons:
            if l.time_start and l.time_end and l.time_start <= now <= l.time_end:
                return jsonify({'lesson': l.to_dict()})
        for l in lessons:
            if l.time_start and l.time_start > now:
                return jsonify({'lesson': l.to_dict()})
        return jsonify({'lesson': lessons[-1].to_dict()})

    @app.route('/api/lessons/<int:lid>')
    @jwt_required()
    def lessons_get(lid):
        return jsonify(Lesson.query.get_or_404(lid).to_dict())

    @app.route('/api/lessons', methods=['POST'])
    @jwt_required()
    def lessons_create():
        """Создаёт занятие. {subject_id*, group_id*, lesson_date*, lesson_number*, ...}"""
        admin_only()
        d = request.get_json(silent=True) or {}
        if not all([d.get('subject_id'), d.get('group_id'),
                    d.get('lesson_date'), d.get('lesson_number')]):
            return err('subject_id, group_id, lesson_date, lesson_number обязательны')
        try: ld = date.fromisoformat(d['lesson_date'])
        except ValueError: return err('Формат даты: YYYY-MM-DD')
        try:
            l = Lesson(subject_id=d['subject_id'], group_id=d['group_id'],
                       teacher_id=d.get('teacher_id'), lesson_number=d['lesson_number'],
                       topic=d.get('topic','').strip() or None, lesson_date=ld,
                       time_start=d.get('time_start','').strip() or None,
                       time_end=d.get('time_end','').strip() or None,
                       classroom=d.get('classroom','').strip() or None)
            db.session.add(l); db.session.commit(); return jsonify(l.to_dict()), 201
        except Exception as e: db.session.rollback(); return err(str(e), 500)

    @app.route('/api/lessons/<int:lid>', methods=['PUT'])
    @jwt_required()
    def lessons_update(lid):
        """Изменяет занятие. Все поля необязательны."""
        admin_only()
        lesson = Lesson.query.get_or_404(lid)
        d = request.get_json(silent=True) or {}
        for f in ('subject_id','group_id','teacher_id','lesson_number',
                  'topic','time_start','time_end','classroom'):
            if f in d:
                v = d[f]
                setattr(lesson, f, v.strip() or None if isinstance(v, str) else v)
        if 'lesson_date' in d:
            try: lesson.lesson_date = date.fromisoformat(d['lesson_date'])
            except ValueError: return err('Формат даты: YYYY-MM-DD')
        try: db.session.commit(); return jsonify(lesson.to_dict())
        except Exception as e: db.session.rollback(); return err(str(e), 500)

    @app.route('/api/lessons/<int:lid>', methods=['DELETE'])
    @jwt_required()
    def lessons_delete(lid):
        admin_only()
        l = Lesson.query.get_or_404(lid)
        try: db.session.delete(l); db.session.commit(); return jsonify({'ok':True})
        except Exception as e: db.session.rollback(); return err(str(e), 500)

    #  Студенты 

    @app.route('/api/students')
    @jwt_required()
    def students_list():
        q = Student.query
        s = request.args.get('q','').strip()
        gid = request.args.get('group_id', type=int)
        if s:
            p = f'%{s}%'
            q = q.filter(db.or_(Student.first_name.ilike(p),
                                  Student.last_name.ilike(p),
                                  Student.student_id.ilike(p)))
        if gid: q = q.filter(Student.group_id == gid)
        return jsonify([s.to_dict() for s in
                        q.order_by(Student.last_name, Student.first_name).all()])

    @app.route('/api/students/<int:sid>')
    @jwt_required()
    def students_get(sid):
        return jsonify(Student.query.get_or_404(sid).to_dict(with_photos=True))

    @app.route('/api/students', methods=['POST'])
    @jwt_required()
    def students_create():
        """Создаёт студента (multipart). Поля: student_id*, first_name*, last_name*, ..., photo?"""
        admin_only()
        num = (request.form.get('student_id') or '').strip()
        fn  = (request.form.get('first_name') or '').strip()
        ln  = (request.form.get('last_name')  or '').strip()
        if not num or not fn or not ln: return err('Номер зачётки, имя и фамилия обязательны')
        if Student.query.filter_by(student_id=num).first(): return err(f'Номер "{num}" занят')
        try:
            s = Student(student_id=num, first_name=fn, last_name=ln,
                        middle_name=(request.form.get('middle_name') or '').strip() or None,
                        group_id=request.form.get('group_id', type=int),
                        email=(request.form.get('email') or '').strip() or None)
            db.session.add(s); db.session.flush()
            pf = request.files.get('photo')
            if pf and pf.filename: _save_photo(s.id, pf, primary=True)
            db.session.commit(); return jsonify(s.to_dict()), 201
        except ValueError as e: db.session.rollback(); return err(str(e))
        except Exception as e:  db.session.rollback(); return err(str(e), 500)

    @app.route('/api/students/<int:sid>', methods=['PUT'])
    @jwt_required()
    def students_update(sid):
        """Изменяет студента. {student_id, first_name, last_name, middle_name, group_id, email, status}"""
        admin_only()
        s = Student.query.get_or_404(sid)
        d = request.get_json(silent=True) or {}
        for f in ('first_name','last_name','middle_name','email','status'):
            if f in d:
                v = d[f]
                setattr(s, f, (v or '').strip() or None if isinstance(v, str) else v)
        if 'group_id' in d:
            s.group_id = int(d['group_id']) if d['group_id'] else None
        if 'student_id' in d and d['student_id']:
            nn = d['student_id'].strip()
            if nn != s.student_id and Student.query.filter_by(student_id=nn).first():
                return err(f'Номер "{nn}" занят')
            s.student_id = nn
        try: db.session.commit(); return jsonify(s.to_dict())
        except Exception as e: db.session.rollback(); return err(str(e), 500)

    @app.route('/api/students/<int:sid>', methods=['DELETE'])
    @jwt_required()
    def students_delete(sid):
        admin_only()
        s = Student.query.get_or_404(sid)
        pdir = os.path.join(config.PHOTOS_DIR, str(sid))
        try:
            if os.path.isdir(pdir): shutil.rmtree(pdir)
            db.session.delete(s); db.session.commit(); return jsonify({'ok':True})
        except Exception as e: db.session.rollback(); return err(str(e), 500)

    @app.route('/api/students/<int:sid>/photos', methods=['POST'])
    @jwt_required()
    def photos_add(sid):
        admin_only()
        s = Student.query.get_or_404(sid)
        pf = request.files.get('photo')
        if not pf or not pf.filename: return err('Файл не передан')
        try:
            _save_photo(sid, pf, primary=len(s.photos) == 0)
            db.session.commit(); return jsonify(s.to_dict(with_photos=True)), 201
        except ValueError as e: db.session.rollback(); return err(str(e))
        except Exception as e:  db.session.rollback(); return err(str(e), 500)

    @app.route('/api/photos/<int:pid>', methods=['DELETE'])
    @jwt_required()
    def photos_delete(pid):
        admin_only()
        p = StudentPhoto.query.get_or_404(pid)
        if p.file_path and os.path.exists(p.file_path):
            try: os.remove(p.file_path)
            except OSError: pass
        try: db.session.delete(p); db.session.commit(); return jsonify({'ok':True})
        except Exception as e: db.session.rollback(); return err(str(e), 500)

    #  Асинхронное распознавание (polling) 

    @app.route('/api/attendance/recognize', methods=['POST'])
    @jwt_required()
    def attendance_recognize():
        """
        Принимает фото, запускает фоновую задачу, сразу возвращает job_id.

        Клиент опрашивает /api/attendance/jobs/<job_id> каждые 3 сек.
        Поля формы: lesson_id*, photos (один или несколько файлов).
        """
        lesson_id = request.form.get('lesson_id', type=int)
        if not lesson_id: return err('lesson_id обязателен')
        lesson = Lesson.query.get(lesson_id)
        if not lesson: return err('Занятие не найдено', 404)

        photos = request.files.getlist('photos')
        if not photos or not photos[0].filename: return err('Необходимо хотя бы одно фото')

        temp_paths = []
        for pf in photos:
            try:
                img = Image.open(pf).convert('RGB')
                tmp = os.path.join(config.TEMP_DIR, f'{uuid.uuid4().hex}.jpg')
                img.save(tmp, 'JPEG', quality=92)
                temp_paths.append(tmp)
            except Exception as e:
                print(f'recognize: не удалось открыть фото: {e}')

        if not temp_paths: return err('Ни одно фото не удалось обработать')

        job_id = uuid.uuid4().hex
        with _jobs_lock:
            _jobs[job_id] = {'status': 'processing'}

        threading.Thread(
            target=_run_recognition,
            args=(flask_app, job_id, lesson_id, temp_paths),
            daemon=True,
        ).start()

        return jsonify({'job_id': job_id, 'status': 'processing'}), 202

    @app.route('/api/attendance/jobs/<job_id>')
    @jwt_required()
    def attendance_job_status(job_id):
        """Статус задачи. Статусы: processing | done | error."""
        with _jobs_lock:
            job = _jobs.get(job_id)
        if not job: return err('Задача не найдена', 404)
        return jsonify(job)

    @app.route('/api/attendance/submit', methods=['POST'])
    @jwt_required()
    def attendance_submit():
        """Сохраняет журнал. {lesson_id, present_ids[], manual_ids[]}."""
        d = request.get_json(silent=True) or {}
        lid = d.get('lesson_id')
        present = set(d.get('present_ids', []))
        manual  = set(d.get('manual_ids',  []))
        if not lid: return err('lesson_id обязателен')
        lesson = Lesson.query.get(lid)
        if not lesson: return err('Занятие не найдено', 404)
        try:
            Attendance.query.filter_by(lesson_id=lid).delete()
            students = Student.query.filter_by(group_id=lesson.group_id, status='active').all()
            for s in students:
                status = 'manual' if s.id in manual else 'present' if s.id in present else 'absent'
                db.session.add(Attendance(lesson_id=lid, student_id=s.id, status=status))
            db.session.commit()
            pc = len(present | manual)
            return jsonify({'ok':True, 'lesson_id':lid, 'present_count':pc,
                            'absent_count': len(students) - pc})
        except Exception as e: db.session.rollback(); return err(str(e), 500)

    @app.route('/api/attendance/lesson/<int:lid>')
    @jwt_required()
    def attendance_for_lesson(lid):
        lesson  = Lesson.query.get_or_404(lid)
        records = Attendance.query.filter_by(lesson_id=lid).order_by(Attendance.status).all()
        return jsonify({'lesson': lesson.to_dict(), 'records': [r.to_dict() for r in records]})

    @app.route('/api/attendance/summary')
    @jwt_required()
    def attendance_summary():
        """Список всех журналов. Query: group_id, subject_id."""
        sub = db.session.query(Attendance.lesson_id).distinct().subquery()
        q   = Lesson.query.filter(Lesson.id.in_(sub))
        gid = request.args.get('group_id',   type=int)
        sid = request.args.get('subject_id', type=int)
        if gid: q = q.filter(Lesson.group_id   == gid)
        if sid: q = q.filter(Lesson.subject_id == sid)
        lessons = q.order_by(Lesson.lesson_date.desc(), Lesson.time_start.desc()).all()
        result = []
        for l in lessons:
            recs    = Attendance.query.filter_by(lesson_id=l.id).all()
            present = sum(1 for r in recs if r.status in ('present','manual'))
            e       = l.to_dict()
            e.update({'present_count': present, 'absent_count': len(recs) - present,
                      'total_count': len(recs)})
            result.append(e)
        return jsonify(result)

    @app.route('/api/attendance/report')
    @jwt_required()
    def attendance_report():
        """Отчёт посещаемости. Query: group_id*, subject_id."""
        gid = request.args.get('group_id',   type=int)
        sid = request.args.get('subject_id', type=int)
        if not gid: return err('group_id обязателен')
        q = Lesson.query.filter_by(group_id=gid)
        if sid: q = q.filter_by(subject_id=sid)
        lessons  = q.order_by(Lesson.lesson_date).all()
        students = Student.query.filter_by(group_id=gid, status='active').order_by(Student.last_name).all()
        report   = []
        for s in students:
            att = sum(1 for l in lessons
                      if (r := Attendance.query.filter_by(lesson_id=l.id, student_id=s.id).first())
                         and r.status in ('present','manual'))
            e = s.to_dict()
            e.update({'attended': att, 'total': len(lessons),
                      'percentage': round(att / len(lessons) * 100, 1) if lessons else 0})
            report.append(e)
        g = Group.query.get(gid)
        sb = Subject.query.get(sid) if sid else None
        return jsonify({'group': g.to_dict() if g else None,
                        'subject': sb.to_dict() if sb else None,
                        'lessons': [l.to_dict() for l in lessons], 'students': report})

    #  Импорт расписания 

    @app.route('/api/schedule/parse', methods=['POST'])
    @jwt_required()
    def schedule_parse():
        """Разбирает xlsx и возвращает превью. Форма: file."""
        admin_only()
        f = request.files.get('file')
        if not f or not f.filename: return err('Файл не передан')
        if not f.filename.lower().endswith(('.xlsx','.xlsm')): return err('Только .xlsx')
        raw = f.read()
        try: preview = parse_schedule_xlsx(BytesIO(raw))
        except Exception as e: return err(f'Ошибка разбора: {e}')
        tmp = f'{uuid.uuid4().hex}_sched.xlsx'
        try:
            with open(os.path.join(config.TEMP_DIR, tmp), 'wb') as fh: fh.write(raw)
        except OSError as e: return err(str(e), 500)
        preview['temp_file'] = tmp
        return jsonify(preview)

    @app.route('/api/schedule/import', methods=['POST'])
    @jwt_required()
    def schedule_import():
        """
        Импортирует занятия из xlsx.
        Тело: {temp_file, mappings: [{file_group, group_id, teacher_id}]}.
        """
        admin_only()
        d = request.get_json(silent=True) or {}
        tmp_file = d.get('temp_file','').strip()
        mappings = d.get('mappings', [])
        if not tmp_file or not mappings: return err('temp_file и mappings обязательны')
        tmp_path = os.path.join(config.TEMP_DIR, tmp_file)
        if not os.path.exists(tmp_path): return err('Временный файл не найден. Загрузите снова.')
        try:
            with open(tmp_path, 'rb') as fh:
                data = parse_schedule_xlsx(BytesIO(fh.read()), full=True)
        except Exception as e: return err(f'Ошибка чтения: {e}')

        group_map = {m['file_group'].strip(): {
            'group_id':   int(m['group_id']),
            'teacher_id': int(m['teacher_id']) if m.get('teacher_id') else None,
        } for m in mappings if m.get('file_group') and m.get('group_id')}

        sc = lc = sk = 0
        sub_cache = {}
        try:
            for rec in data.get('records', []):
                mapping = group_map.get(rec.get('group_name',''))
                if not mapping: sk += 1; continue
                name = (rec.get('subject') or '').strip()
                if not name or len(name) < 3: sk += 1; continue
                if name not in sub_cache:
                    ex = Subject.query.filter(Subject.name.ilike(name)).first()
                    if ex: sub_cache[name] = ex.id
                    else:
                        ns = Subject(name=name); db.session.add(ns); db.session.flush()
                        sub_cache[name] = ns.id; sc += 1
                if Lesson.query.filter_by(
                    subject_id=sub_cache[name], group_id=mapping['group_id'],
                    lesson_date=rec['date'], lesson_number=rec['lesson_number']
                ).first(): sk += 1; continue
                db.session.add(Lesson(
                    subject_id=sub_cache[name], group_id=mapping['group_id'],
                    teacher_id=mapping['teacher_id'], lesson_number=rec['lesson_number'],
                    lesson_date=rec['date'], time_start=rec.get('time_start'),
                    time_end=rec.get('time_end'), classroom=rec.get('classroom'),
                ))
                lc += 1
            db.session.commit()
        except Exception as e: db.session.rollback(); return err(f'Ошибка импорта: {e}', 500)
        try: os.remove(tmp_path)
        except OSError: pass
        return jsonify({'subjects_created': sc, 'lessons_created': lc, 'skipped': sk})

    #  Настройки 

    @app.route('/api/settings')
    @jwt_required()
    def settings_get():
        return jsonify(get_service().get_settings())

    @app.route('/api/settings', methods=['PUT'])
    @jwt_required()
    def settings_update():
        admin_only()
        d = request.get_json(silent=True) or {}
        if 'threshold' in d:
            try:
                t = float(d['threshold'])
                if not 0.1 <= t <= 0.9: return err('threshold: 0.1–0.9')
                get_service().update_threshold(t)
            except (TypeError, ValueError): return err('threshold должен быть числом')
        return jsonify(get_service().get_settings())


#  Фоновая задача 

def _run_recognition(flask_app, job_id, lesson_id, temp_paths):
    """
    Распознавание лиц в фоновом потоке.

    Использует flask_app.app_context() — единственный способ работать
    с SQLAlchemy и Flask-расширениями вне HTTP-запроса.
    """
    with flask_app.app_context():
        try:
            lesson  = Lesson.query.get(lesson_id)
            service = get_service()

            if not lesson:
                with _jobs_lock: _jobs[job_id] = {'status':'error','message':'Занятие не найдено'}
                return

            group_photos = StudentPhoto.query.join(Student).filter(
                Student.group_id == lesson.group_id,
                Student.status   == 'active',
                StudentPhoto.embedding  != None,
                StudentPhoto.model_name == service.model,
            ).all()

            if not group_photos:
                with _jobs_lock:
                    _jobs[job_id] = {'status':'error',
                                     'message': f'Нет эмбеддингов для группы "{lesson.group.name}". '
                                                 'Загрузите фотографии студентов.'}
                return

            candidates = []
            for p in group_photos:
                try: candidates.append((p.student_id, service.deserialize_embedding(p.embedding)))
                except Exception as e: print(f'_run_recognition: десериализация photo_id={p.id}: {e}')

            recognized_map    = {}
            unrecognized_count = 0

            for tmp in temp_paths:
                try:
                    for emb in service.extract_all_faces_embeddings(tmp):
                        match = service.find_best_match(emb, candidates)
                        if match['matched']:
                            sid = match['student_id']
                            if sid not in recognized_map or match['confidence'] > recognized_map[sid]['confidence']:
                                recognized_map[sid] = match
                        else:
                            unrecognized_count += 1
                except Exception as e:
                    print(f'_run_recognition: ошибка {tmp}: {e}')
                finally:
                    if os.path.exists(tmp):
                        try: os.remove(tmp)
                        except OSError: pass

            all_students = Student.query.filter_by(
                group_id=lesson.group_id, status='active').order_by(Student.last_name).all()

            recognized = []
            for sid, m in recognized_map.items():
                s = Student.query.get(sid)
                if s:
                    e = s.to_dict(); e['confidence'] = m['confidence']; e['distance'] = m['distance']
                    recognized.append(e)

            absent = [s.to_dict() for s in all_students if s.id not in recognized_map]

            with _jobs_lock:
                _jobs[job_id] = {'status': 'done', 'result': {
                    'lesson': lesson.to_dict(), 'recognized': recognized,
                    'absent': absent, 'unrecognized_count': unrecognized_count,
                    'photos_processed': len(temp_paths),
                }}

        except Exception as e:
            import traceback; traceback.print_exc()
            for tmp in temp_paths:
                if os.path.exists(tmp):
                    try: os.remove(tmp)
                    except OSError: pass
            with _jobs_lock:
                _jobs[job_id] = {'status': 'error', 'message': str(e)}


# Парсер Excel 

def parse_schedule_xlsx(file_obj, full=False):
    """
    Разбирает xlsx-расписание университетского формата.

    Файл содержит до 124 листов (один лист = одна неделя).
    8 групп в столбцах 3–10, 6 дней в неделю.

    Дедупликация групп по первому токену названия:
    «ОБ-Вт-09.03.03.02-11 ВТ 1 курс (15 чел)» и
    «ОБ-Вт-09.03.03.02-11 (15 чел) 1 курс» → один ключ «ОБ-Вт-09.03.03.02-11».
    Это гарантирует ровно 8 уникальных групп независимо от числа листов.
    """
    from openpyxl import load_workbook
    from datetime import datetime as dt_type

    ROMAN    = {'I':1,'II':2,'III':3,'IV':4,'V':5,'VI':6,'VII':7}
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    records  = []
    all_dates  = []
    groups_map = {}   # code -> full_name (первое вхождение)

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 8:
            continue

        # Строка-заголовок (содержит «дни» или «Время»)
        hi = None
        for i, row in enumerate(rows[:12]):
            if row and any(str(c).strip().lower() in ('дни','время') for c in row if c):
                hi = i; break
        if hi is None:
            continue

        group_row = rows[hi + 1] if hi + 1 < len(rows) else []

        # Столбцы 3–10: коды групп.
        # Ключ дедупликации = первый токен (до пробела).
        col_groups = {}
        for ci in range(3, min(13, len(group_row))):
            v = group_row[ci]
            if not v or not str(v).strip():
                continue
            full = re.sub(r'\s+', ' ', str(v)).strip()
            tokens = full.split()
            if not tokens:
                continue
            code = tokens[0]   # e.g. «ОБ-Вт-09.03.03.02-11»
            col_groups[ci] = code
            if code not in groups_map:
                groups_map[code] = full

        if not col_groups:
            continue

        current_date = None
        i = hi + 2

        while i < len(rows):
            row = rows[i]
            if not row or not any(c for c in row):
                i += 1; continue

            # Столбец 0: дата
            c0 = row[0] if row else None
            if c0:
                if isinstance(c0, dt_type):
                    current_date = c0.date(); all_dates.append(current_date)
                elif isinstance(c0, date):
                    current_date = c0; all_dates.append(current_date)
                elif isinstance(c0, str):
                    m = re.search(r'(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})', c0)
                    if m:
                        d2, mo, y = m.groups(); y = int(y)
                        if y < 100: y += 2000
                        try: current_date = date(y, int(mo), int(d2)); all_dates.append(current_date)
                        except ValueError: pass

            # Столбец 1: номер пары (I–VII)
            c1 = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            if c1 not in ROMAN: i += 1; continue

            lesson_number = ROMAN[c1]

            # Столбец 2: время «8.20-09.50»
            c2 = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            tm = re.search(r'(\d{1,2})[.:h](\d{2})\s*[-–]\s*(\d{1,2})[.:h](\d{2})', c2)
            ts = f'{int(tm.group(1)):02d}:{tm.group(2)}' if tm else None
            te = f'{int(tm.group(3)):02d}:{tm.group(4)}' if tm else None

            if not current_date: i += 1; continue

            next_row = rows[i + 1] if i + 1 < len(rows) else []

            for ci, group_code in col_groups.items():
                subject = classroom = teacher = None
                if ci < len(row) and row[ci]:
                    raw = re.sub(r'\s+', ' ', str(row[ci])).strip()
                    if len(raw) >= 4 and 'поздравл' not in raw.lower():
                        subject = raw
                if next_row and ci < len(next_row) and next_row[ci]:
                    raw = re.sub(r'\s+', ' ', str(next_row[ci])).strip().split('\n')[0].strip()
                    rm  = re.search(r'(\d+\s*/\s*\d+)', raw)
                    if rm:
                        classroom = rm.group(1).replace(' ','')
                        teacher   = re.sub(r'[,.\s]+$', '', raw[:rm.start()].strip()) or None
                    else:
                        teacher = re.sub(r'https?://\S+', '', raw).strip() or None

                if subject:
                    if full:
                        records.append({'date': current_date, 'lesson_number': lesson_number,
                                        'time_start': ts, 'time_end': te,
                                        'group_name': group_code, 'subject': subject,
                                        'teacher': teacher, 'classroom': classroom})
                    else:
                        records.append(True)
            i += 1

    return {
        'groups':       sorted(groups_map.keys()),
        'lesson_count': len(records),
        'date_range': {'from': str(min(all_dates)) if all_dates else None,
                       'to':   str(max(all_dates)) if all_dates else None},
        **(({'records': records}) if full else {}),
    }


# Утилиты

def err(message, code=400):
    return jsonify({'error': message}), code

def admin_only():
    if get_jwt().get('role') != 'admin': abort(403)

def _save_photo(student_id, file_obj, primary=False):
    """Сохраняет фото, извлекает эмбеддинг, создаёт StudentPhoto."""
    ext = file_obj.filename.rsplit('.', 1)[-1].lower() if '.' in file_obj.filename else ''
    if ext not in config.ALLOWED_IMAGE_EXTENSIONS:
        raise ValueError('Поддерживаются только JPEG и PNG')
    folder = os.path.join(config.PHOTOS_DIR, str(student_id))
    os.makedirs(folder, exist_ok=True)
    filename = f'{uuid.uuid4().hex}.{ext}'
    filepath = os.path.join(folder, filename)
    file_obj.save(filepath)
    service = get_service()
    emb = service.extract_embedding(filepath)
    if emb is None:
        try: os.remove(filepath)
        except OSError: pass
        raise ValueError('Лицо не обнаружено на фотографии')
    if primary:
        StudentPhoto.query.filter_by(student_id=student_id, is_primary=True).update({'is_primary': False})
    db.session.add(StudentPhoto(
        student_id=student_id, filename=filename, file_path=filepath,
        embedding=service.serialize_embedding(emb),
        model_name=service.model, detector_backend=service.detector, is_primary=primary,
    ))
    db.session.flush()

def _seed_users():
    if User.query.count() > 0: return
    for login, pwd, role, name in [
        (config.DEFAULT_ADMIN_LOGIN,   config.DEFAULT_ADMIN_PASSWORD,   'admin',   'Администратор'),
        (config.DEFAULT_TEACHER_LOGIN, config.DEFAULT_TEACHER_PASSWORD, 'teacher', 'Преподаватель'),
    ]:
        u = User(username=login, role=role, full_name=name); u.set_password(pwd); db.session.add(u)
    db.session.commit()
    print(f'Учётные записи: {config.DEFAULT_ADMIN_LOGIN}/{config.DEFAULT_ADMIN_PASSWORD}'
          f'  |  {config.DEFAULT_TEACHER_LOGIN}/{config.DEFAULT_TEACHER_PASSWORD}')

if __name__ == '__main__':
    application = create_app()
    print('Система учёта посещаемости: http://localhost:5000')
    application.run(host='0.0.0.0', port=5000, debug=False, ssl_context=('cert.pem', 'key.pem'))
